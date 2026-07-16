"""Pure Excel builder (SPEC §10). Reload the bytes with openpyxl and assert the sheet — no I/O."""

from io import BytesIO

from openpyxl import load_workbook

from app.utils.excel import orders_workbook

_ROW = {
    "order_number": 7, "customer_name": "Sara", "phone": "+9715", "address": "Marina",
    "quantity": 2, "selling_price": "6000", "discount_amount": "400", "status": "confirmed",
    "delivery_date": "2026-07-11", "created_at": "2026-07-10T14:30:22+00:00",
    "special_instructions": "call first",
    "products": {"category": "Mobile", "brand": "Samsung", "model": "S23 Ultra",
                 "color": "Green", "specs": {"RAM": "12GB", "storage": "256GB"}},
    "delivery_persons": {"name": "Ali", "phone": "+9716"},
}


def _sheet(rows, **kw):
    ws = load_workbook(BytesIO(orders_workbook(rows, **kw))).active
    hdr = [c.value for c in ws[1]]
    return ws, hdr


def test_headers_and_style():
    ws, hdr = _sheet([])
    assert hdr[:4] == ["Order ID", "Customer Name", "Phone", "Address"]
    assert hdr[-1] == "Status" and "RAM" in hdr and "Storage" in hdr
    assert ws["A1"].fill.fgColor.rgb.endswith("2563EB")  # SPEC §10 blue header
    assert ws["A1"].font.bold and ws["A1"].font.color.rgb.endswith("FFFFFF")


def test_row_mapping_and_net_price():
    ws, hdr = _sheet([_ROW])
    r = {h: c.value for h, c in zip(hdr, ws[2])}
    assert r["Order ID"] == 7 and r["Color"] == "Green"
    assert r["RAM"] == "12GB" and r["Storage"] == "256GB"  # extracted from specs (case-insensitive)
    assert r["Selling Price"] == 5600  # 6000 charged - 400 discount = what the customer pays


def test_detailed_adds_rider_columns():
    ws, hdr = _sheet([_ROW], detailed=True)
    assert {"Order Time", "Rider Name", "Rider Phone", "Special Instructions"} <= set(hdr)
    r = {h: c.value for h, c in zip(hdr, ws[2])}
    assert r["Rider Name"] == "Ali" and r["Special Instructions"] == "call first"
    assert r["Order Time"].startswith("2026-07-10 14:30")  # ISO T → space


def test_base_has_no_detail_columns():
    _, hdr = _sheet([_ROW])
    assert "Rider Name" not in hdr and "Order Time" not in hdr


# --- generic sheet builder (the counter sheet reuses the orders styling) ---
def test_sheet_workbook_styles_header_and_writes_rows():
    from io import BytesIO

    from openpyxl import load_workbook

    from app.utils.excel import sheet_workbook

    data = sheet_workbook("Counter sales", ["A", "B"], [[1, "x"], [2, "y"]])
    ws = load_workbook(BytesIO(data)).active
    assert ws.title == "Counter sales"
    assert [c.value for c in ws[1]] == ["A", "B"]
    assert ws["A1"].fill.fgColor.rgb.endswith("2563EB")  # same SPEC §10 header as orders
    assert ws.freeze_panes == "A2"
    assert [c.value for c in ws[2]] == [1, "x"]


def test_counter_sheet_rows_leave_sale_columns_empty_for_the_pen():
    from app.products.service import counter_sheet_rows

    rows = counter_sheet_rows([
        {"product_number": 1, "brand": "Samsung", "model": "S23", "color": "black",
         "specs": {"ram": "12GB", "storage": "256GB"}, "quantity": 4},
        {"product_number": 2, "brand": "Apple", "model": "iPhone 15", "quantity": 0},
    ])
    assert rows[0][:4] == ["PR0001", "Samsung S23 (black)", "ram: 12GB · storage: 256GB", 4]
    assert rows[0][4:] == ["", ""]          # Price sold / Qty sold — the shop writes these by hand
    assert rows[1][:2] == ["PR0002", "Apple iPhone 15"]
    assert rows[1][2] == ""                 # no specs → empty cell, not "None"
