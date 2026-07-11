"""Pick-&-Pack Excel builder (SPEC §10). Pure: order rows → .xlsx bytes, no I/O.

Input is the DB order shape used everywhere else — each row carries an embedded
`products` dict (and, for rider sheets, `delivery_persons`). Keeping the flatten +
RAM/Storage-from-specs extraction here (not in the service) makes the whole mapping
unit-testable with hand-built dicts and no database.
"""

from __future__ import annotations

from io import BytesIO
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

# SPEC §10: white bold on blue #2563EB, thin borders, auto-width.
_HEADER_FILL = PatternFill("solid", fgColor="2563EB")
_HEADER_FONT = Font(bold=True, color="FFFFFF")
_THIN = Side(style="thin", color="D0D0D0")
_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)
_CENTER = Alignment(horizontal="center", vertical="center")

_BASE = [
    "Order ID", "Customer Name", "Phone", "Address", "Product", "Brand", "Model",
    "Color", "RAM", "Storage", "Qty", "Selling Price", "Delivery Date", "Status",
]
_DETAIL = ["Order Time", "Rider Name", "Rider Phone", "Special Instructions"]  # SPEC §10 detailed


def _spec(specs: dict | None, *keys: str) -> str:
    """First present spec value by lowercase key (specs keys are 'ram', 'storage', ...)."""
    low = {str(k).lower(): v for k, v in (specs or {}).items()}
    for k in keys:
        v = low.get(k)
        if v not in (None, ""):
            return str(v)
    return ""


def _row(o: dict, detailed: bool) -> list[Any]:
    p = o.get("products") or {}
    net = float(o.get("selling_price") or 0) - float(o.get("discount_amount") or 0)  # what the customer pays
    row = [
        o.get("order_number"), o.get("customer_name"), o.get("phone"), o.get("address"),
        p.get("category"), p.get("brand"), p.get("model"), p.get("color"),
        _spec(p.get("specs"), "ram"), _spec(p.get("specs"), "storage", "rom"),
        o.get("quantity"), net, o.get("delivery_date"), o.get("status"),
    ]
    if detailed:
        dp = o.get("delivery_persons") or {}
        row += [
            (o.get("created_at") or "")[:19].replace("T", " "),  # ISO → 'YYYY-MM-DD HH:MM:SS'
            dp.get("name"), dp.get("phone"), o.get("special_instructions"),
        ]
    return row


def orders_workbook(rows: list[dict], *, detailed: bool = False) -> bytes:
    """Build the pick-&-pack sheet. `detailed` adds time/rider/instructions columns (SPEC §10)."""
    headers = _BASE + (_DETAIL if detailed else [])
    wb = Workbook()
    ws = wb.active
    ws.title = "Orders"

    ws.append(headers)
    for c in ws[1]:
        c.fill, c.font, c.border, c.alignment = _HEADER_FILL, _HEADER_FONT, _BORDER, _CENTER

    for o in rows:
        ws.append(_row(o, detailed))
    for line in ws.iter_rows(min_row=2):
        for c in line:
            c.border = _BORDER

    for i in range(1, len(headers) + 1):
        letter = ws.cell(row=1, column=i).column_letter
        longest = max((len(str(c.value)) for c in ws[letter] if c.value is not None), default=10)
        ws.column_dimensions[letter].width = min(max(longest + 2, 10), 50)  # auto-width, capped
    ws.freeze_panes = "A2"  # header stays visible while packing a long sheet

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


if __name__ == "__main__":  # ponytail self-check: builds, reloads, columns + specs extraction hold
    from openpyxl import load_workbook

    sample = [{
        "order_number": 7, "customer_name": "Sara", "phone": "+9715", "address": "Marina",
        "quantity": 2, "selling_price": "6000", "discount_amount": "400", "status": "confirmed",
        "delivery_date": "2026-07-11", "created_at": "2026-07-10T14:30:22+00:00",
        "special_instructions": "call first",
        "products": {"category": "Mobile", "brand": "Samsung", "model": "S23 Ultra",
                     "color": "Green", "specs": {"RAM": "12GB", "storage": "256GB"}},
        "delivery_persons": {"name": "Ali", "phone": "+9716"},
    }]
    wb = load_workbook(BytesIO(orders_workbook(sample, detailed=True)))
    ws = wb.active
    hdr = [c.value for c in ws[1]]
    assert hdr[:4] == ["Order ID", "Customer Name", "Phone", "Address"], hdr
    assert "Rider Name" in hdr and "Order Time" in hdr, hdr
    assert ws["A1"].fill.fgColor.rgb.endswith("2563EB"), ws["A1"].fill.fgColor.rgb
    r = {h: c.value for h, c in zip(hdr, ws[2])}
    assert r["RAM"] == "12GB" and r["Storage"] == "256GB", r
    assert r["Selling Price"] == 5600, r["Selling Price"]  # 6000 - 400
    assert r["Rider Name"] == "Ali" and r["Order Time"].startswith("2026-07-10 14:30"), r
    print("excel.py self-check OK")
